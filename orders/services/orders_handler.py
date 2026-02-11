from io import BytesIO
from datetime import datetime
from django.db import transaction
from django.core.exceptions import ValidationError
from rest_framework import serializers
from openpyxl import load_workbook
from orders.models import Order, OrderProduct, StockMovement
from django.db import transaction
from django.contrib.auth import get_user_model

from products.models import Product, UnitOfMeasure

User = get_user_model()




SHIPPING_COST = 8000


VALID_ORDER_STATUS = {
    "PENDING", "PROCESSING", "SHIPPED", "OUT_FOR_DELIVERY",
    "DELIVERED", "CANCELLED", "RETURNED", "FAILED", "ON_HOLD"
}

VALID_DISCOUNT_TYPES = {
    "REFERRAL", "FIRST_PURCHASE", "COUPON", "SEASONAL", "NONE"
}


def validate_excel_data(orders_sheet, items_sheet):
    errors = []

    # =========================
    # ORDERS VALIDATION
    # =========================
    order_ids = set()

    for row_number, (
        order_id,
        user_dni,
        status,
        discount_applied,
        discount_type,
        discount_value,
        shipping_cost,
    ) in enumerate(
        orders_sheet.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        # order_id (can be empty → auto-generated)
        if order_id:
            order_ids.add(order_id)

        # user_dni
        if not user_dni or not str(user_dni).isdigit():
            errors.append(f"Orders row {row_number}: invalid user_dni")

        # status
        if status not in VALID_ORDER_STATUS:
            errors.append(
                f"Orders row {row_number}: invalid status '{status}'"
            )

        # discount_applied
        if discount_applied not in (True, False, "TRUE", "FALSE"):
            errors.append(
                f"Orders row {row_number}: invalid discount_applied value"
            )

        # normalize discount_applied
        discount_applied_bool = discount_applied in (True, "TRUE")

        # discount_type
        if discount_type not in VALID_DISCOUNT_TYPES:
            errors.append(
                f"Orders row {row_number}: invalid discount_type '{discount_type}'"
            )

        # discount consistency
        if not discount_applied_bool and discount_value > 0:
            errors.append(
                f"Orders row {row_number}: discount_value > 0 but discount_applied is FALSE"
            )

        if discount_value < 0:
            errors.append(
                f"Orders row {row_number}: discount_value cannot be negative"
            )

        if shipping_cost < 0:
            errors.append(
                f"Orders row {row_number}: shipping_cost cannot be negative"
            )

    # =========================
    # ORDER PRODUCTS VALIDATION
    # =========================
    for row_number, (
        order_id,
        product_id,
        price,
        quantity,
        measure_unit_id,
    ) in enumerate(
        items_sheet.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        # FK to Order
        if order_id not in order_ids:
            errors.append(
                f"OrderProducts row {row_number}: order_id '{order_id}' does not exist in Orders sheet"
            )

        # quantity
        if not isinstance(quantity, (int, float)) or quantity <= 0:
            errors.append(
                f"OrderProducts row {row_number}: invalid quantity"
            )

        # price
        if not isinstance(price, (int, float)) or price < 0:
            errors.append(
                f"OrderProducts row {row_number}: invalid price"
            )

    if errors:
        raise ValidationError(errors)

class OrdersFileSerializer(serializers.Serializer):
    file = serializers.FileField()

class UploadOrderFileSerializer(serializers.Serializer):
    file = serializers.FileField()

class OrdersFileParser:
    """
    Parses an Excel (.xlsx) file and returns a list of product dictionaries
    ready to be validated by a DRF serializer.
    """

    def parse(self, file) -> tuple:
        """
        file: InMemoryUploadedFile | TemporaryUploadedFile
        """

        workbook = load_workbook(
            filename=BytesIO(file.read()),
            read_only=True,
            data_only=True
        )

        orders_sheet = workbook["Orders"]
        order_items_sheet = workbook["OrderProduct"]

        return orders_sheet, order_items_sheet


        return orders, items




User = get_user_model()


class OrdersUploadFileService:

    def execute(self, orders_sheet, items_sheet) -> dict:
        with transaction.atomic():
            orders_map = self._create_or_update_orders(orders_sheet)
            items_count = self._create_order_items(items_sheet, orders_map)

        #TODO: Create a StockMovement register with the current OrderProducts

        stock_movement = StockMovement()




        return {
            "orders": len(orders_map),
            "items": items_count,
        }

    def _create_or_update_orders(self, orders_sheet) -> dict:
        orders_map = {}

        for row in orders_sheet.iter_rows(min_row=2, values_only=True):
            # Filtrar filas vacías
            if not any(row):
                continue
                
            try:
                order_id, user_dni, status, discount_applied, discount_type, discount_value, shipping_cost = row
            except ValueError as e:
                # Debug: ver qué fila está causando el problema
                print(f"Error en fila: {row}")
                continue

            try:
                user_dni = int(user_dni)
            except (ValueError, TypeError):
                print(f"DNI inválido: {user_dni}")
                continue


            try:
                user = User.objects.get(dni=str(user_dni))
            except User.DoesNotExist as e:
                return {'message': f'user not found with DNI: {user_dni}'}

            discount_applied_bool = discount_applied in (True, "TRUE", "True", "true")

            order_data = {
                "user": user,
                "status": status,
                "discount_applied": discount_applied_bool,
                "discount_type": discount_type,
                "discount_value": float(discount_value) if discount_value else 0.0,
                "shipping_cost": float(shipping_cost) if shipping_cost else 0.0,
            }

            if order_id:
                order, _ = Order.objects.update_or_create(
                    id=order_id,
                    defaults=order_data,
                )
            else:
                order = Order.objects.create(**order_data)

            orders_map[order.id] = order

        return orders_map


    def _create_order_items(self, items_sheet, orders_map) -> int:
        order_products = []

        product_cache = {p.sku: p for p in Product.objects.all()}
        unit_cache = {u.id: u for u in UnitOfMeasure.objects.all()}

        for row in items_sheet.iter_rows(min_row=2, values_only=True):
            # Saltar filas vacías
            if not any(cell for cell in row[:5] if cell not in (None, '')):
                continue
                
            if len(row) < 5:
                continue
                
            order_id, product_sku, price, quantity, measure_unit_id = row[:5]
            
            # Validaciones básicas
            if not all([order_id, product_sku, price, quantity]):
                continue
                
            if order_id not in orders_map:
                continue
                
            if product_sku not in product_cache:
                continue

            # Preparar datos
            order = orders_map[order_id]
            product = product_cache[product_sku]
            
            # Manejar unidad de medida
            unit = None
            if measure_unit_id:
                try:
                    unit_id = int(measure_unit_id)
                    unit = unit_cache.get(unit_id)
                except:
                    pass

            try:
                price = float(price)
                quantity = int(quantity)
            except:
                continue

            order_products.append(
                OrderProduct(
                    order=order,
                    product=product,
                    price=price,
                    quantity=quantity,
                    measure_unity=unit,
                )
            )

        if order_products:
            OrderProduct.objects.bulk_create(order_products)
            
        return len(order_products)
